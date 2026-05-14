from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from app.models import InventoryItem, InventoryMovement


def stock_status(item: InventoryItem) -> str:
    if item.current_stock <= item.min_stock:
        return 'Stock bajo'
    if item.max_stock > 0 and item.current_stock >= item.max_stock:
        return 'Sobre stock'
    return 'Normal'


def build_inventory_excel(items: list[InventoryItem]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    headers = ['Código', 'Artículo', 'Descripción', 'Unidad', 'Stock actual', 'Stock mínimo', 'Stock máximo', 'Estado', 'Actualizado']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1E293B')
        cell.alignment = Alignment(horizontal='center')

    for item in items:
        ws.append([
            item.code, item.name, item.description, item.unit,
            item.current_stock, item.min_stock, item.max_stock,
            stock_status(item), item.updated_at.strftime('%Y-%m-%d %H:%M') if item.updated_at else '',
        ])

    widths = [16, 28, 42, 14, 16, 16, 16, 18, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_movements_excel(movements: list[InventoryMovement]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimientos'
    headers = ['Fecha', 'Código', 'Artículo', 'Tipo', 'Cantidad', 'Stock antes', 'Stock después', 'Motivo', 'Notas', 'Realizado por']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1E293B')
        cell.alignment = Alignment(horizontal='center')

    for mov in movements:
        ws.append([
            mov.created_at.strftime('%Y-%m-%d %H:%M') if mov.created_at else '',
            mov.item_code,
            mov.item_name,
            mov.movement_type,
            mov.quantity,
            mov.stock_before,
            mov.stock_after,
            mov.reason,
            mov.notes,
            mov.performed_by_name,
        ])

    widths = [20, 16, 30, 14, 12, 14, 16, 30, 35, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_inventory_pdf(items: list[InventoryItem]) -> BytesIO:
    stream = BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph('Reporte de Inventario de Mantenimiento', styles['Title']), Spacer(1, 12)]

    data = [['Código', 'Artículo', 'Unidad', 'Stock', 'Mínimo', 'Máximo', 'Estado']]
    for item in items:
        data.append([item.code, item.name, item.unit, item.current_stock, item.min_stock, item.max_stock, stock_status(item)])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(table)
    doc.build(story)
    stream.seek(0)
    return stream


def build_movements_pdf(movements: list[InventoryMovement]) -> BytesIO:
    stream = BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph('Reporte de Movimientos de Inventario', styles['Title']), Spacer(1, 12)]

    data = [['Fecha', 'Código', 'Artículo', 'Tipo', 'Cantidad', 'Stock', 'Motivo', 'Realizado por']]
    for mov in movements:
        data.append([
            mov.created_at.strftime('%Y-%m-%d %H:%M') if mov.created_at else '',
            mov.item_code,
            mov.item_name,
            mov.movement_type,
            mov.quantity,
            f'{mov.stock_before} → {mov.stock_after}',
            mov.reason,
            mov.performed_by_name,
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(table)
    doc.build(story)
    stream.seek(0)
    return stream
